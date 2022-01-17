import openpyxl
# 엑셀을 쉽게 다룰 수 있도록 도와주는 라이브러리 

# 엑셀 만들기
wb = openpyxl.Workbook()
# 엑셀 시트 만들기
ws = wb.create_sheet('오징어 게임')
# 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 엑셀 저장하기 
wb.save('참가자_data.xlsx')
