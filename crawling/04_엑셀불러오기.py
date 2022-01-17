import openpyxl

url = "/Users/choesuyeon/Desktop/crawling/참가자_data.xlsx"
# 엑셀 불러오기 : 역슬랙시 일 경우 앞에 r(문자열로 취급해라)--> 윈도우 
wb = openpyxl.load_workbook(url)

# 엑셀 시트 선택 
ws = wb['오징어 게임']
# 데이터 수정하기 
ws['A3'] = 456
ws['B3'] = '성기훈'

# 엑셀 저장하기
wb.save(url)